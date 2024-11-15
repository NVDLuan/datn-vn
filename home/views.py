from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib import messages
from django.http import JsonResponse
from datetime import date
from .models import PhongLuuTru, DonVi, VanBan, HoSo, HoSo, PhongLuuTru, PhanLoaiTaiLieuLuuTru, ThoiHanLuuTru, MucLucHoSo, HopLuuTru
from .forms import PhongLuuTruForm, VanBanForm, HoSoForm
from django.contrib.auth import views as auth_views
from .actions import print_to_docx
import openpyxl
from django.http import HttpResponse
from datetime import datetime
from django.contrib.auth.views import LoginView, PasswordResetView, PasswordChangeView, PasswordResetConfirmView
from admin_tabler.forms import RegistrationForm, LoginForm, UserPasswordResetForm, UserSetPasswordForm, UserPasswordChangeForm
from django.views.decorators.csrf import csrf_protect
def print_docx_view(request, app_label, model_name, pk):
    return print_to_docx(request, app_label, model_name, pk)
# Custom LoginView
# class CustomLoginView(auth_views.LoginView):
#     template_name = 'pages/sign-in.html'
#     form_class = LoginForm
#     redirect_authenticated_user = True  # Add if you want to redirect already logged-in users

# class LoginView(LoginView):
#     template_name = 'pages/sign-in.html'
#     form_class = LoginForm
#     redirect_authenticated_user = True  # Add if you want to redirect already logged-in users
  
@login_required
def index(request):
    # Page from the theme 
    return render(request, 'pages/index.html')

@login_required
def hoso(request):
    # Page from the theme 
    return render(request, 'pages/hoso.html')

@login_required
def phongluutru(request):
    """
    Hiển thị danh sách phòng lưu trữ.

    Args:
        request: Django request object.

    Returns:
        HttpResponse: Rendered HTML template.
    """
    # Lấy danh sách đơn vị mà user có quyền truy cập
    don_vi_ids = request.user.userdonvi_set.values_list('don_vi', flat=True)
    phongluutru_list = PhongLuuTru.objects.filter(don_vi__in=don_vi_ids).order_by('-id')

    # Tìm kiếm
    query = request.GET.get('q')
    if query:
        phongluutru_list = phongluutru_list.filter(
            Q(name__icontains=query) | Q(phong_so__icontains=query)
        )

    # Phân trang
    paginator = Paginator(phongluutru_list, 10) # 10 bản ghi mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'pages/phongluutru.html', context)

@login_required
def phongluutru_themmoi(request):
    """
    Tạo mới phòng lưu trữ.

    Args:
        request: Django request object.

    Returns:
        HttpResponse: Rendered HTML template hoặc JsonResponse.
    """
    if request.method == 'POST':
        form = PhongLuuTruForm(request.POST)
        if form.is_valid():
            phongluutru = form.save(commit=False)
            phongluutru.don_vi = request.user.userdonvi_set.first().don_vi
            phongluutru.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Thêm mới phông lưu trữ thành công!'})
            else:
                messages.success(request, 'Thêm mới phông lưu trữ thành công!')
                return redirect('phongluutru')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors})
            else:
                messages.error(request, 'Lỗi khi thêm mới phông lưu trữ. Vui lòng kiểm tra lại.')
    else:
        form = PhongLuuTruForm()

    context = {'form': form}
    return render(request, 'pages/phongluutru_themmoi.html', context)

@login_required
def phongluutru_chitiet(request, id):
    """
    Hiển thị chi tiết phòng lưu trữ.

    Args:
        request: Django request object.
        id: ID của phòng lưu trữ.

    Returns:
        HttpResponse: Rendered HTML template.
    """
    phongluutru = get_object_or_404(PhongLuuTru, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if phongluutru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     messages.error(request, 'Bạn không có quyền truy cập phông lưu trữ này.')
    #     return redirect('phongluutru')

    context = {'phongluutru': phongluutru}
    return render(request, 'pages/phongluutru_chitiet.html', context)

@login_required
def phongluutru_capnhat(request, id):
    """
    Cập nhật phòng lưu trữ.

    Args:
        request: Django request object.
        id: ID của phòng lưu trữ.

    Returns:
        HttpResponse: Rendered HTML template hoặc JsonResponse.
    """
    phongluutru = get_object_or_404(PhongLuuTru, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if phongluutru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     messages.error(request, 'Bạn không có quyền truy cập phông lưu trữ này.')
    #     return redirect('phongluutru')

    if request.method == 'POST':
        form = PhongLuuTruForm(request.POST, instance=phongluutru)
        if form.is_valid():
            form.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Cập nhật phông lưu trữ thành công!'})
            else:
                messages.success(request, 'Cập nhật phông lưu trữ thành công!')
                return redirect('phongluutru')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors})
            else:
                messages.error(request, 'Lỗi khi cập nhật phông lưu trữ. Vui lòng kiểm tra lại.')
    else:
        form = PhongLuuTruForm(instance=phongluutru)

    context = {'form': form, 'phongluutru': phongluutru}
    return render(request, 'pages/phongluutru_capnhat.html', context)

@login_required
def phongluutru_xoa(request, id):
    """
    Xóa phòng lưu trữ.

    Args:
        request: Django request object.
        id: ID của phòng lưu trữ.

    Returns:
        JsonResponse: JSON response cho biết kết quả của thao tác.
    """
    phongluutru = get_object_or_404(PhongLuuTru, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if phongluutru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     return JsonResponse({'status': 'error', 'message': 'Bạn không có quyền truy cập phông lưu trữ này.'})

    if request.method == 'POST':
        try:
            phongluutru.delete()
            return JsonResponse({'status': 'success', 'message': 'Xóa phông lưu trữ thành công!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Lỗi khi xóa phông lưu trữ: {e}'})

    return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'})


@login_required
def vanban(request):
    """
    Hiển thị danh sách văn bản.

    Args:
        request: Django request object.

    Returns:
        HttpResponse: Rendered HTML template.
    """
    # Lấy danh sách Hồ sơ mà user có quyền truy cập
    ho_so_ids = HoSo.objects.filter(
        phong_luu_tru__don_vi__in=request.user.userdonvi_set.values_list('don_vi', flat=True)
    ).values_list('id', flat=True)
    vanban_list = VanBan.objects.filter(ho_so__in=ho_so_ids).order_by('-id')

    # Tìm kiếm
    query = request.GET.get('q')
    if query:
        vanban_list = vanban_list.filter(
            Q(so_van_ban__icontains=query) | Q(ky_hieu_cua_van_ban__icontains=query)
        )

    # Phân trang
    paginator = Paginator(vanban_list, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'pages/vanban.html', context)

@login_required
def vanban_themmoi(request):
    """
    Tạo mới văn bản.

    Args:
        request: Django request object.

    Returns:
        HttpResponse: Rendered HTML template hoặc JsonResponse.
    """
    if request.method == 'POST':
        form = VanBanForm(request.POST)
        if form.is_valid():
            form.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Thêm mới văn bản thành công!'})
            else:
                messages.success(request, 'Thêm mới văn bản thành công!')
                return redirect('vanban')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors})
            else:
                messages.error(request, 'Lỗi khi thêm mới văn bản. Vui lòng kiểm tra lại.')
    else:
        form = VanBanForm()

    context = {'form': form}
    return render(request, 'pages/vanban_themmoi.html', context)

@login_required
def vanban_chitiet(request, id):
    """
    Hiển thị chi tiết văn bản.

    Args:
        request: Django request object.
        id: ID của văn bản.

    Returns:
        HttpResponse: Rendered HTML template.
    """
    vanban = get_object_or_404(VanBan, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if vanban.ho_so.phong_luu_tru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     messages.error(request, 'Bạn không có quyền truy cập văn bản này.')
    #     return redirect('vanban')

    context = {'vanban': vanban}
    return render(request, 'pages/vanban_chitiet.html', context)

@login_required
def vanban_capnhat(request, id):
    """
    Cập nhật văn bản.

    Args:
        request: Django request object.
        id: ID của văn bản.

    Returns:
        HttpResponse: Rendered HTML template hoặc JsonResponse.
    """
    vanban = get_object_or_404(VanBan, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if vanban.ho_so.phong_luu_tru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     messages.error(request, 'Bạn không có quyền truy cập văn bản này.')
    #     return redirect('vanban')

    if request.method == 'POST':
        form = VanBanForm(request.POST, instance=vanban)
        if form.is_valid():
            form.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Cập nhật văn bản thành công!'})
            else:
                messages.success(request, 'Cập nhật văn bản thành công!')
                return redirect('vanban')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors})
            else:
                messages.error(request, 'Lỗi khi cập nhật văn bản. Vui lòng kiểm tra lại.')
    else:
        form = VanBanForm(instance=vanban)

    context = {'form': form, 'vanban': vanban}
    return render(request, 'pages/vanban_capnhat.html', context)

@login_required
def vanban_xoa(request, id):
    """
    Xóa văn bản.

    Args:
        request: Django request object.
        id: ID của văn bản.

    Returns:
        JsonResponse: JSON response cho biết kết quả của thao tác.
    """
    vanban = get_object_or_404(VanBan, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if vanban.ho_so.phong_luu_tru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     return JsonResponse({'status': 'error', 'message': 'Bạn không có quyền truy cập văn bản này.'})

    if request.method == 'POST':
        try:
            vanban.delete()
            return JsonResponse({'status': 'success', 'message': 'Xóa văn bản thành công!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Lỗi khi xóa văn bản: {e}'})

    return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'})


@login_required
def hoso(request):
    """
    Hiển thị danh sách hồ sơ.

    Args:
        request: Django request object.

    Returns:
        HttpResponse: Rendered HTML template.
    """
    # Lấy danh sách PhongLuuTru mà user có quyền truy cập
    phong_luu_tru_ids = PhongLuuTru.objects.filter(
        don_vi__in=request.user.userdonvi_set.values_list('don_vi', flat=True)
    ).values_list('id', flat=True)
    try:
        hoso_list = HoSo.objects.filter(phong_luu_tru__in=phong_luu_tru_ids).order_by('-id')
        # Sắp xếp bằng Python, chuyển chuỗi thành số trước khi sắp xếp
        hoso_list = sorted(hoso_list, key=lambda obj: int(obj.ho_so_so))


        # Tìm kiếm
        query = request.GET.get('q')
        if query:
            hoso_list = hoso_list.filter(
                Q(tieu_de__icontains=query) | Q(ho_so_so__icontains=query)
            )
        else:
            # Sắp xếp bằng Python, chuyển chuỗi thành số trước khi sắp xếp
            hoso_list = sorted(hoso_list, key=lambda obj: int(obj.ho_so_so))

        # Phân trang
        paginator = Paginator(hoso_list, 10) 
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    except:
        hoso_list = HoSo.objects.filter(phong_luu_tru__in=phong_luu_tru_ids).order_by('-id')

        # Tìm kiếm
        query = request.GET.get('q')
        if query:
            hoso_list = hoso_list.filter(
                Q(tieu_de__icontains=query) | Q(ho_so_so__icontains=query)
            )
        else:
            # Sắp xếp bằng Python, chuyển chuỗi thành số trước khi sắp xếp
            hoso_list = sorted(hoso_list, key=lambda obj: int(obj.ho_so_so))

        # Phân trang
        paginator = Paginator(hoso_list, 10) 
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'pages/hoso.html', context)

@login_required
def hoso_themmoi(request):
    """
    Tạo mới hồ sơ.

    Args:
        request: Django request object.

    Returns:
        HttpResponse: Rendered HTML template hoặc JsonResponse.
    """
    if request.method == 'POST':
        form = HoSoForm(request.POST)
        if form.is_valid():
            form.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Thêm mới hồ sơ thành công!'})
            else:
                messages.success(request, 'Thêm mới hồ sơ thành công!')
                return redirect('hoso')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors})
            else:
                messages.error(request, 'Lỗi khi thêm mới hồ sơ. Vui lòng kiểm tra lại.')
    else:
        form = HoSoForm()

    context = {'form': form}
    return render(request, 'pages/hoso_themmoi.html', context)

@login_required
def hoso_chitiet(request, id):
    """
    Hiển thị chi tiết hồ sơ.

    Args:
        request: Django request object.
        id: ID của hồ sơ.

    Returns:
        HttpResponse: Rendered HTML template.
    """
    hoso = get_object_or_404(HoSo, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if hoso.phong_luu_tru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     messages.error(request, 'Bạn không có quyền truy cập hồ sơ này.')
    #     return redirect('hoso')

    context = {'hoso': hoso}
    return render(request, 'pages/hoso_chitiet.html', context)

@login_required
def hoso_capnhat(request, id):
    """
    Cập nhật hồ sơ.

    Args:
        request: Django request object.
        id: ID của hồ sơ.

    Returns:
        HttpResponse: Rendered HTML template hoặc JsonResponse.
    """
    hoso = get_object_or_404(HoSo, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if hoso.phong_luu_tru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     messages.error(request, 'Bạn không có quyền truy cập hồ sơ này.')
    #     return redirect('hoso')

    if request.method == 'POST':
        form = HoSoForm(request.POST, instance=hoso)
        if form.is_valid():
            form.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Cập nhật hồ sơ thành công!'})
            else:
                messages.success(request, 'Cập nhật hồ sơ thành công!')
                return redirect('hoso')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors})
            else:
                messages.error(request, 'Lỗi khi cập nhật hồ sơ. Vui lòng kiểm tra lại.')
    else:
        form = HoSoForm(instance=hoso)

    context = {'form': form, 'hoso': hoso}
    return render(request, 'pages/hoso_capnhat.html', context)

@login_required
def hoso_xoa(request, id):
    """
    Xóa hồ sơ.

    Args:
        request: Django request object.
        id: ID của hồ sơ.

    Returns:
        JsonResponse: JSON response cho biết kết quả của thao tác.
    """
    hoso = get_object_or_404(HoSo, id=id)

    # # Kiểm tra quyền truy cập dựa trên đơn vị
    # if hoso.phong_luu_tru.don_vi not in request.user.userdonvi_set.values_list('don_vi', flat=True):
    #     return JsonResponse({'status': 'error', 'message': 'Bạn không có quyền truy cập hồ sơ này.'})

    if request.method == 'POST':
        try:
            hoso.delete()
            return JsonResponse({'status': 'success', 'message': 'Xóa hồ sơ thành công!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Lỗi khi xóa hồ sơ: {e}'})

    return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ.'})

import json
from django.shortcuts import render
from django.db.models import Count, F, ExpressionWrapper, fields
from django.db.models.functions import ExtractYear, ExtractDay
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import JsonResponse
from .models import VanBan, HoSo, DonVi, PhongLuuTru

@login_required
def thong_ke_van_ban(request):
    # Lấy đơn vị thuộc quyền user
    don_vi_ids = request.user.userdonvi_set.values_list('don_vi', flat=True)

    # Thống kê số lượng văn bản đã số hóa
    so_hoa_theo_ngay = (
        VanBan.objects.filter(ho_so__phong_luu_tru__don_vi__in=don_vi_ids)
        .annotate(ngay=F("ngay_ban_hanh"))
        .values("ngay")
        .annotate(so_luong=Count("id"))
        .order_by("ngay")
    )

    so_hoa_theo_thang = (
        VanBan.objects.filter(ho_so__phong_luu_tru__don_vi__in=don_vi_ids)
        .annotate(thang=F("ngay_ban_hanh__month"), nam=F("ngay_ban_hanh__year"))
        .values("thang", "nam")
        .annotate(so_luong=Count("id"))
        .order_by("nam", "thang")
    )

    so_hoa_theo_quy = (
        VanBan.objects.filter(ho_so__phong_luu_tru__don_vi__in=don_vi_ids)
        .annotate(
            quy=F("ngay_ban_hanh__quarter"), nam=F("ngay_ban_hanh__year")
        )
        .values("quy", "nam")
        .annotate(so_luong=Count("id"))
        .order_by("nam", "quy")
    )

    so_hoa_theo_nam = (
        VanBan.objects.filter(ho_so__phong_luu_tru__don_vi__in=don_vi_ids)
        .annotate(nam=F("ngay_ban_hanh__year"))
        .values("nam")
        .annotate(so_luong=Count("id"))
        .order_by("nam")
    )
    # Chuyển đổi date sang string trước khi serialize
    so_hoa_theo_ngay = list(so_hoa_theo_ngay)
    for item in so_hoa_theo_ngay:
        item['ngay'] = item['ngay'].strftime('%Y-%m-%d')

    so_hoa_theo_thang = list(so_hoa_theo_thang)
    for item in so_hoa_theo_thang:
        item['thang'] = item['thang'].strftime('%Y-%m') if hasattr(item['thang'], 'strftime') else item['thang']

    so_hoa_theo_quy = list(so_hoa_theo_quy)
    for item in so_hoa_theo_quy:
        item['quy'] = f"Quý {item['quy']}/{item['nam']}" 

    so_hoa_theo_nam = list(so_hoa_theo_nam)
    for item in so_hoa_theo_nam:
        item['nam'] = item['nam'].strftime('%Y') if hasattr(item['nam'], 'strftime') else item['nam']

    # Thống kê số lượng văn bản
    tong_so_van_ban = VanBan.objects.filter(
        ho_so__phong_luu_tru__don_vi__in=don_vi_ids
    ).count()
    so_van_ban_da_so_hoa = VanBan.objects.filter(
        trang_thai='da_so_hoa', ho_so__phong_luu_tru__don_vi__in=don_vi_ids
    ).count()
    so_van_ban_chua_so_hoa = tong_so_van_ban - so_van_ban_da_so_hoa

    # Tính tỷ lệ phần trăm
    ty_le_da_so_hoa = (so_van_ban_da_so_hoa / tong_so_van_ban) * 100 if tong_so_van_ban > 0 else 0
    ty_le_chua_so_hoa = (so_van_ban_chua_so_hoa / tong_so_van_ban) * 100 if tong_so_van_ban > 0 else 0
    print(tong_so_van_ban)
    print(ty_le_da_so_hoa)
    print(ty_le_chua_so_hoa)

    context = {
        'tong_so_van_ban': tong_so_van_ban,
        'ty_le_da_so_hoa': round(ty_le_da_so_hoa),
        'ty_le_chua_so_hoa': round(ty_le_chua_so_hoa),
        'so_luong_van_ban': {  # Dữ liệu cho biểu đồ mới
            'da_so_hoa': so_van_ban_da_so_hoa,
            'chua_so_hoa': so_van_ban_chua_so_hoa,
            'tong_so': tong_so_van_ban
        },
        'so_hoa_theo_ngay': json.dumps(list(so_hoa_theo_ngay)),
        'so_hoa_theo_thang': json.dumps(list(so_hoa_theo_thang)),
        'so_hoa_theo_quy': json.dumps(list(so_hoa_theo_quy)),
        'so_hoa_theo_nam': json.dumps(list(so_hoa_theo_nam)),
        # "ho_so_het_han": json.dumps(list(ho_so_het_han)),
    }
    return render(request, "pages/thong_ke_van_ban.html", context)


@login_required
def thong_ke_don_vi_ajax(request):
    don_vi_id = request.GET.get("don_vi")

    if not don_vi_id:
        return JsonResponse({"error": "Thiếu tham số đơn vị"}, status=400)

    try:
        don_vi = DonVi.get_cached_don_vi(don_vi_id)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)

    # Lấy danh sách phòng lưu trữ thuộc đơn vị
    phong_luu_tru_ids = PhongLuuTru.objects.filter(
        don_vi=don_vi
    ).values_list("id", flat=True)

    # Thống kê số lượng văn bản đã số hóa theo đơn vị
    so_hoa_theo_don_vi = (
        VanBan.objects.filter(ho_so__phong_luu_tru__in=phong_luu_tru_ids)
        .annotate(ngay=F("ngay_ban_hanh"))
        .values("ngay")
        .annotate(so_luong=Count("id"))
        .order_by("ngay")
    )

    return JsonResponse({"data": list(so_hoa_theo_don_vi)})

def hoso_hethan_view(request):
    from datetime import timedelta
    today = date.today()
    user = request.user

    # Lấy đơn vị thuộc quyền user
    don_vi_ids = request.user.userdonvi_set.values_list('don_vi', flat=True)

    # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và có thời hạn bảo quản
    ho_so_co_thoi_han = HoSo.objects.filter(
        thoi_gian_ket_thuc__lt=today, 
        phong_luu_tru__don_vi__in=don_vi_ids
    ).exclude(thoi_han_bao_quan__ten="Vĩnh viễn")

    # Lấy ngày hiện tại
    now = datetime.now().date()

    # Lọc danh sách hồ sơ hết hạn
    ho_so_het_han = []
    for ho_so in ho_so_co_thoi_han:
        thoi_gian_het_han = ho_so.thoi_gian_bat_dau + timedelta(days=ho_so.thoi_han_bao_quan.so_nam * 365.25)
        if thoi_gian_het_han < now:
            ho_so_het_han.append(ho_so)


    total_hoso_hethan = len(ho_so_het_han)
    context = {
        'hoso_hethan_list': ho_so_het_han,
        'total_hoso_hethan': total_hoso_hethan,
    }
    return render(request, 'pages/thong_ke_ho_so_het_han.html', context)


from django.conf import settings
import os

@login_required
def export_mucluchoso_excel(request, ho_so_id):
    ho_so = get_object_or_404(HoSo, pk=ho_so_id)
    phong_luu_tru = ho_so.phong_luu_tru
    don_vi = phong_luu_tru.don_vi
    muc_luc_ho_so = get_object_or_404(MucLucHoSo, pk=ho_so.muc_luc_ho_so.id)

    # Lấy danh sách hồ sơ cùng mục lục hồ sơ, phông và đơn vị
    ho_so_list = HoSo.objects.filter(
        muc_luc_ho_so=ho_so.muc_luc_ho_so,
        phong_luu_tru=phong_luu_tru,
        phong_luu_tru__don_vi=don_vi
    )

    template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)

    # Load mẫu Excel
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Điền tên đơn vị
    worksheet['A1'] = f"{don_vi.ten_don_vi}"

    # Điền ngày tháng năm hiện tại
    worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"


    # Điền thông tin Mục lục hồ sơ
    worksheet['A3'] = f"MỤC LỤC HỒ SƠ ({muc_luc_ho_so.muc_luc_so})"
    
    # Điền dữ liệu Hồ sơ
    row_num = 6  # Bắt đầu từ dòng 6
    for hoso in ho_so_list:
        worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
        worksheet['B{}'.format(row_num)] = hoso.ho_so_so
        worksheet['C{}'.format(row_num)] = hoso.tieu_de
        worksheet['D{}'.format(row_num)] = f"{hoso.thoi_gian_bat_dau.strftime('%d/%m/%Y')} - {hoso.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if hoso.thoi_gian_ket_thuc else ''}"
        worksheet['E{}'.format(row_num)] = f"{hoso.so_luong_to}"
        worksheet['F{}'.format(row_num)] = hoso.ghi_chu
        row_num += 1

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="muc_luc_ho_so_{muc_luc_ho_so.muc_luc_so}.xlsx"'
    workbook.save(response)
    return response

@login_required
def export_hopluutru_excel(request, ho_so_id):
    ho_so = get_object_or_404(HoSo, pk=ho_so_id)
    phong_luu_tru = ho_so.phong_luu_tru
    don_vi = phong_luu_tru.don_vi
    kho_luu_tru = ho_so.phong_luu_tru.kho_luu_tru   
    hop_luu_tru = get_object_or_404(HopLuuTru, pk=ho_so.hop_luu_tru.id)
    # muc_luc_ho_so = get_object_or_404(MucLucHoSo, pk=ho_so.muc_luc_ho_so.id)

    # Lấy danh sách hồ sơ cùng mục lục hồ sơ, phông và đơn vị
    ho_so_list = HoSo.objects.filter(
        hop_luu_tru=ho_so.hop_luu_tru,
        phong_luu_tru=phong_luu_tru,
        phong_luu_tru__don_vi=don_vi
    )

    # Lấy hồ sơ số nhỏ nhất và lớn nhất
    ho_so_so_min = ho_so_list.order_by('ho_so_so').first().ho_so_so if ho_so_list else ''
    ho_so_so_max = ho_so_list.order_by('-ho_so_so').first().ho_so_so if ho_so_list else ''

    # Lấy năm hồ sơ nhỏ nhất và lớn nhất
    nam_ho_so_min = ho_so_list.order_by('thoi_gian_bat_dau').first().thoi_gian_bat_dau.year if ho_so_list else ''
    nam_ho_so_max = ho_so_list.order_by('-thoi_gian_bat_dau').first().thoi_gian_bat_dau.year if ho_so_list else ''

    # Số lượng ĐVBQ (Đơn vị văn bản quản lý)
    so_luong_dvbq = ho_so_list.count()

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '6.1.mauhop-ngang.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Điền dữ liệu
    worksheet['A1'] = f"KHO LƯU TRỮ\n{kho_luu_tru.ten_kho}"
    worksheet['A2'] = f"PHÔNG LƯU TRỮ\n{phong_luu_tru.name}"
    worksheet['A4'] = f"HỘP SỐ"
    worksheet['A5'] = hop_luu_tru.hop_so
    worksheet['A6'] = f"Gồm: {so_luong_dvbq} ĐVBQ"
    worksheet['A7'] = f"Hồ sơ số: {ho_so_so_min}-{ho_so_so_max}"
    worksheet['A14'] = f"NĂM"
    if nam_ho_so_min == nam_ho_so_max:
        worksheet['A15'] = f"{nam_ho_so_min}"
    elif nam_ho_so_min == "":
        worksheet['A15'] = f"{nam_ho_so_max}"
    elif nam_ho_so_max == "":
        worksheet['A15'] = f"{nam_ho_so_min}"
    else:
        worksheet['A15'] = f"{nam_ho_so_min} - {nam_ho_so_max}"

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="hop_so_{hop_luu_tru.hop_so}.xlsx"'
    workbook.save(response)
    return response

@login_required
def export_muclucvanban_excel(request, hoso_id):
    ho_so = get_object_or_404(HoSo, pk=hoso_id)
    van_ban_list = VanBan.objects.filter(ho_so=ho_so)

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '5.MucLucVanBan-(6).xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Điền ngày tháng năm
    worksheet['E3'] = f"Đà nẵng, ngày {ngay} tháng {thang} năm {nam}"

     # Điền tên đơn vị
    worksheet['A1'] = f"{ho_so.phong_luu_tru.don_vi.ten_don_vi}"

    # Điền thông tin Hồ sơ
    worksheet['A4'] = f"MỤC LỤC VĂN BẢN (HỒ SƠ {ho_so.ho_so_so})"
    # Điền thông tin Hồ sơ
    worksheet['A5'] = f"Hồ sơ : {ho_so.tieu_de}"

    # Điền dữ liệu Văn bản
    row_num = 9  # Bắt đầu từ dòng 9
    for idx, van_ban in enumerate(van_ban_list):
        worksheet['A{}'.format(row_num)] = idx+1
        worksheet['B{}'.format(row_num)] = f"{van_ban.so_van_ban}/{van_ban.ky_hieu_cua_van_ban}"
        worksheet['C{}'.format(row_num)] = van_ban.ngay_ban_hanh.strftime('%d/%m/%Y')
        worksheet['D{}'.format(row_num)] = f"{van_ban.loai_van_ban} - {van_ban.trich_yeu_noi_dung}" if van_ban.loai_van_ban else van_ban.trich_yeu_noi_dung 
        worksheet['E{}'.format(row_num)] = van_ban.co_quan_to_chuc_ban_hanh
        worksheet['F{}'.format(row_num)] = f"{van_ban.tu_so}-{van_ban.den_so}" if van_ban.tu_so and van_ban.den_so else ''
        worksheet['G{}'.format(row_num)] = van_ban.ghi_chu
        row_num += 1

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="muc_luc_van_ban_{ho_so.ho_so_so}.xlsx"'
    workbook.save(response)
    return response

@login_required
def export_biahoso_excel(request, hoso_id):
    ho_so = get_object_or_404(HoSo, pk=hoso_id)

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '7.bia.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Điền dữ liệu
    worksheet['B3'] = f"{str(ho_so.phong_luu_tru.don_vi.ten_don_vi).upper()}"
    worksheet['B4'] = f"{ho_so.phong_luu_tru.name}"  # Giả sử tên trung tâm luôn cố định
    worksheet['C9'] = ho_so.ma_ho_so
    # worksheet['A14'] = "HỒ SƠ"
    worksheet['A15'] = ho_so.tieu_de
    worksheet['D18'] = f"Từ ngày: {ho_so.thoi_gian_bat_dau.strftime('%d/%m/%Y')}"
    worksheet['F18'] = f"Đến ngày: {ho_so.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if ho_so.thoi_gian_ket_thuc else ''}"
    worksheet['B19'] = f"Gồm: {ho_so.so_luong_to} tờ"
    # worksheet['C23'] = "Phông số:"
    worksheet['C22'] = ho_so.phong_luu_tru.phong_so
    # worksheet['C24'] = "Mục lục số:"
    worksheet['C23'] = ho_so.muc_luc_ho_so.muc_luc_so if ho_so.muc_luc_ho_so else ''
    # worksheet['C25'] = "Hồ sơ số:"
    worksheet['C24'] = ho_so.ho_so_so
    # worksheet['F23'] = "THỜI HẠN BẢO QUẢN"
    # worksheet['F23'] = f"{ho_so.thoi_han_bao_quan.so_nam} NĂM" if ho_so.thoi_han_bao_quan else ''
    worksheet['G23'] = ho_so.thoi_han_bao_quan.ten if ho_so.thoi_han_bao_quan else ''

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="In_bia_ho_so_{ho_so.ho_so_so}.xlsx"'
    workbook.save(response)
    return response

@login_required
def export_bao_cao_thong_ke(request):
    # Lấy đơn vị thuộc quyền user
    don_vi_ids = request.user.userdonvi_set.values_list("don_vi", flat=True)

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'thongke', '8.bao cao thong ke.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Điền ngày tháng năm hiện tại
    worksheet['C22'] = f"Đà nẵng, ngày {ngay} tháng {thang} năm {nam}"

    # Thống kê dữ liệu
    tong_so_phong = PhongLuuTru.objects.filter(don_vi__in=don_vi_ids).count()
    tong_so_mlhs = MucLucHoSo.objects.filter(phong_luu_tru__don_vi__in=don_vi_ids).count()
    tong_so_ho_so = HoSo.objects.filter(phong_luu_tru__don_vi__in=don_vi_ids).count()
    tong_so_van_ban = VanBan.objects.filter(ho_so__phong_luu_tru__don_vi__in=don_vi_ids).count()
    tong_so_van_ban_da_so_hoa = VanBan.objects.filter(trang_thai='da_so_hoa', ho_so__phong_luu_tru__don_vi__in=don_vi_ids).count()

    thoi_han_bao_quan_list = ThoiHanLuuTru.objects.all()
    thong_ke_ho_so_theo_thoi_han = {}
    for thoi_han in thoi_han_bao_quan_list:
        so_luong = HoSo.objects.filter(thoi_han_bao_quan=thoi_han, phong_luu_tru__don_vi__in=don_vi_ids).count()
        thong_ke_ho_so_theo_thoi_han[thoi_han.ten] = so_luong

    # Điền dữ liệu thống kê vào Excel
    worksheet['C5'] = tong_so_phong
    worksheet['C6'] = tong_so_mlhs
    worksheet['C7'] = tong_so_ho_so
    worksheet['C8'] = tong_so_van_ban
    worksheet['C9'] = tong_so_van_ban_da_so_hoa

    # Thống kê hồ sơ theo thời hạn lưu trữ
    thoi_han_bao_quan_list = ThoiHanLuuTru.objects.all()
    row_num = 10
    for thoi_han in thoi_han_bao_quan_list:
        worksheet['B{}'.format(row_num)] = thoi_han.ten  # Điền tên thời hạn vào cột B
        so_luong = HoSo.objects.filter(
            thoi_han_bao_quan=thoi_han, 
            phong_luu_tru__don_vi__in=don_vi_ids
        ).count()
        worksheet['C{}'.format(row_num)] = so_luong  # Điền số lượng vào cột C
        row_num += 1

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bao_cao_thong_ke_tai_lieu_toan_bo_don_vi.xlsx"'
    workbook.save(response)
    return response

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, F
from django.utils import timezone
from datetime import datetime
import openpyxl
import os
from django.conf import settings

from .models import (
    PhongLuuTru, MucLucHoSo, HoSo, VanBan, ThoiHanLuuTru, DonVi
)

@login_required
def in_va_thong_ke(request):
    # Lấy đơn vị thuộc quyền user
    don_vi_ids = request.user.userdonvi_set.values_list("don_vi", flat=True)

    # Thống kê dữ liệu
    tong_so_phong = PhongLuuTru.objects.filter(don_vi__in=don_vi_ids).count()
    tong_so_mlhs = MucLucHoSo.objects.filter(
        phong_luu_tru__don_vi__in=don_vi_ids
    ).count()
    tong_so_ho_so = HoSo.objects.filter(
        phong_luu_tru__don_vi__in=don_vi_ids
    ).count()
    tong_so_van_ban = VanBan.objects.filter(
        ho_so__phong_luu_tru__don_vi__in=don_vi_ids
    ).count()
    tong_so_van_ban_da_so_hoa = VanBan.objects.filter(
        trang_thai="da_so_hoa", ho_so__phong_luu_tru__don_vi__in=don_vi_ids
    ).count()

    # Thống kê hồ sơ theo thời hạn bảo quản
    thoi_han_bao_quan_list = ThoiHanLuuTru.objects.all()
    thong_ke_ho_so_theo_thoi_han = {}
    for thoi_han in thoi_han_bao_quan_list:
        so_luong = HoSo.objects.filter(
            thoi_han_bao_quan=thoi_han, phong_luu_tru__don_vi__in=don_vi_ids
        ).count()
        thong_ke_ho_so_theo_thoi_han[thoi_han.ten] = so_luong

    context = {
        "tong_so_phong": tong_so_phong,
        "tong_so_mlhs": tong_so_mlhs,
        "tong_so_ho_so": tong_so_ho_so,
        "tong_so_van_ban": tong_so_van_ban,
        "tong_so_van_ban_da_so_hoa": tong_so_van_ban_da_so_hoa,
        "thoi_han_bao_quan_list": thoi_han_bao_quan_list,
        "thong_ke_ho_so_theo_thoi_han": thong_ke_ho_so_theo_thoi_han,
        "don_vi_list": DonVi.objects.filter(id__in=don_vi_ids),
    }
    return render(request, "pages/in_va_thong_ke.html", context)


# In ấn

def in_an_helper(request, template_path, filename, data, sheet_name=None, data_rows=None):
    # Load mẫu Excel
    full_template_path = os.path.join(settings.BASE_DIR, 'home', 'templates', template_path)
    workbook = openpyxl.load_workbook(full_template_path)
    worksheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Điền dữ liệu
    for key, value in data.items():
        worksheet[key] = value.format(ngay=ngay, thang=thang, nam=nam)

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response

def in_thoi_han_bao_quan(request):
    don_vi_id = request.GET.get("don_vi_id")
    thoi_han_id = request.GET.get("thoi_han_id")
    if not thoi_han_id or not don_vi_id:
        return JsonResponse({"error": "Thiếu tham số ID thời hạn"}, status=400)

    try:
        thoi_han = ThoiHanLuuTru.objects.get(pk=thoi_han_id)
        don_vi = DonVi.objects.get(pk=don_vi_id)
        ho_so_list = HoSo.objects.filter(
            thoi_han_bao_quan=thoi_han, phong_luu_tru__don_vi=don_vi
        )
    except ThoiHanLuuTru.DoesNotExist:
        return JsonResponse({"error": "Thời hạn bảo quản không tồn tại"}, status=404)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)

    if thoi_han.ten == "Vĩnh viễn":
        # Lấy thời hạn bảo quản "Vĩnh viễn"
        thoi_han_vinh_vien = ThoiHanLuuTru.objects.get(ten="Vĩnh viễn")

        # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và có thời hạn bảo quản "Vĩnh viễn"
        ho_so_list = HoSo.objects.filter(
            phong_luu_tru__don_vi=don_vi,
            thoi_han_bao_quan=thoi_han_vinh_vien
        )

        # Lấy danh sách tất cả Mục lục hồ sơ có chứa Hồ sơ "Vĩnh viễn"
        muc_luc_ho_so_list = MucLucHoSo.objects.filter(
            id__in=ho_so_list.values_list('muc_luc_ho_so', flat=True).distinct()
        )

        # Load mẫu Excel
        template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs.xlsx')
        full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
        workbook = openpyxl.load_workbook(full_template_path)

        # Lấy sheet đầu tiên làm sheet mẫu
        source_sheet = workbook.active

        # Lấy ngày tháng năm hiện tại
        now = datetime.now()
        ngay = now.day
        thang = now.month
        nam = now.year

        # Duyệt qua danh sách Mục lục hồ sơ
        for i, muc_luc_ho_so in enumerate(muc_luc_ho_so_list):
            if i > 0:
                # Copy sheet mẫu và đổi tên
                target_sheet = workbook.copy_worksheet(source_sheet)
                target_sheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"
                worksheet = target_sheet
            else:
                # Sử dụng sheet đầu tiên
                worksheet = source_sheet
                worksheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"

            # Điền tên đơn vị
            worksheet['A1'] = f"{don_vi.ten_don_vi}"

            # Điền ngày tháng năm hiện tại
            worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"

            # Điền thông tin Mục lục hồ sơ
            worksheet['A3'] = f"MỤC LỤC HỒ SƠ ({muc_luc_ho_so.muc_luc_so})"

            # Điền dữ liệu Hồ sơ
            row_num = 6  # Bắt đầu từ dòng 6
            for hoso in ho_so_list.filter(muc_luc_ho_so=muc_luc_ho_so):
                worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
                worksheet['B{}'.format(row_num)] = hoso.ho_so_so
                worksheet['C{}'.format(row_num)] = hoso.tieu_de
                worksheet['D{}'.format(row_num)] = f"{hoso.thoi_gian_bat_dau.strftime('%d/%m/%Y')} - {hoso.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if hoso.thoi_gian_ket_thuc else ''}"
                worksheet['E{}'.format(row_num)] = f"{hoso.so_luong_to}"
                worksheet['F{}'.format(row_num)] = hoso.ghi_chu
                row_num += 1

        # Xóa sheet mẫu nếu nó vẫn còn sau khi duyệt
        if len(workbook.sheetnames) > len(muc_luc_ho_so_list):
            workbook.remove(source_sheet)

        # Tạo response HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="In_MLHS_theo_thoi_han_vv.xlsx"'
        workbook.save(response)
        return response  
    else:
        # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và KHÔNG có thời hạn bảo quản "Vĩnh viễn"
        ho_so_list = HoSo.objects.filter(
            phong_luu_tru__don_vi=don_vi,
            thoi_han_bao_quan = thoi_han_id
        )

        # Lấy danh sách tất cả Mục lục hồ sơ có chứa Hồ sơ "Có thời hạn"
        muc_luc_ho_so_list = MucLucHoSo.objects.filter(
            id__in=ho_so_list.values_list('muc_luc_ho_so', flat=True).distinct()
        )

        # Load mẫu Excel
        template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs-cth.xlsx') # Thay đổi tên template
        full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
        workbook = openpyxl.load_workbook(full_template_path)

        # Lấy sheet đầu tiên làm sheet mẫu
        source_sheet = workbook.active

        # Lấy ngày tháng năm hiện tại
        now = datetime.now()
        ngay = now.day
        thang = now.month
        nam = now.year

        # Duyệt qua danh sách Mục lục hồ sơ
        for i, muc_luc_ho_so in enumerate(muc_luc_ho_so_list):
            if i > 0:
                # Copy sheet mẫu và đổi tên
                target_sheet = workbook.copy_worksheet(source_sheet)
                target_sheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"
                worksheet = target_sheet
            else:
                # Sử dụng sheet đầu tiên
                worksheet = source_sheet
                worksheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"

            # Điền tên đơn vị
            worksheet['A1'] = f"{don_vi.ten_don_vi}"

            # Điền ngày tháng năm hiện tại
            worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"

            # Điền thông tin Mục lục hồ sơ
            worksheet['A3'] = f"MỤC LỤC HỒ SƠ ({muc_luc_ho_so.muc_luc_so})"

            # Điền dữ liệu Hồ sơ
            row_num = 6  # Bắt đầu từ dòng 6
            for hoso in ho_so_list.filter(muc_luc_ho_so=muc_luc_ho_so):
                worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
                worksheet['B{}'.format(row_num)] = hoso.ho_so_so
                worksheet['C{}'.format(row_num)] = hoso.tieu_de
                worksheet['D{}'.format(row_num)] = f"{hoso.thoi_gian_bat_dau.strftime('%d/%m/%Y')} - {hoso.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if hoso.thoi_gian_ket_thuc else ''}"
                worksheet['E{}'.format(row_num)] = f"{hoso.so_luong_to}"
                # Thêm cột Thời hạn bảo quản
                worksheet['F{}'.format(row_num)] = hoso.thoi_han_bao_quan.ten 
                worksheet['G{}'.format(row_num)] = hoso.ghi_chu
                row_num += 1

        # Xóa sheet mẫu nếu nó vẫn còn sau khi duyệt
        if len(workbook.sheetnames) > len(muc_luc_ho_so_list):
            workbook.remove(source_sheet)

        # Tạo response HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="In_MLHS_theo_thoi_han_bao_quan.xlsx"'
        workbook.save(response)
        return response   

def in_muc_luc_ho_so(request):
    don_vi_id = request.GET.get("don_vi_id")
    muc_luc_ho_so_id = request.GET.get("muc_luc_ho_so_id")
    if not muc_luc_ho_so_id or not don_vi_id:
        return JsonResponse({"error": "Thiếu tham số ID mục lục hồ sơ"}, status=400)

    try:
        muc_luc_ho_so = MucLucHoSo.objects.get(pk=muc_luc_ho_so_id)
        don_vi = DonVi.objects.get(pk=don_vi_id)
        # Lấy danh sách Hồ sơ thuộc Mục lục hồ sơ và Đơn vị
        ho_so_list = HoSo.objects.filter(
            muc_luc_ho_so=muc_luc_ho_so, phong_luu_tru__don_vi=don_vi
        )
    except MucLucHoSo.DoesNotExist:
        return JsonResponse({"error": "Mục lục hồ sơ không tồn tại"}, status=404)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)


    # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và KHÔNG có thời hạn bảo quản "Vĩnh viễn"
    ho_so_list = HoSo.objects.filter(
        phong_luu_tru__don_vi=don_vi,
        muc_luc_ho_so = muc_luc_ho_so_id
    )

    # Lấy danh sách tất cả Mục lục hồ sơ có chứa Hồ sơ "Có thời hạn"
    muc_luc_ho_so_list = MucLucHoSo.objects.filter(
        id__in=ho_so_list.values_list('muc_luc_ho_so', flat=True).distinct()
    )



    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Duyệt qua danh sách Mục lục hồ sơ
    for i, muc_luc_ho_so in enumerate(muc_luc_ho_so_list):
        if ho_so_list[0].thoi_han_bao_quan.ten == "Vĩnh viễn":
            template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs.xlsx') # Thay đổi tên template
        else:
            template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs-cth.xlsx') # Thay đổi tên template
        full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
        workbook = openpyxl.load_workbook(full_template_path)

        # Lấy sheet đầu tiên làm sheet mẫu
        source_sheet = workbook.active

        if i > 0:
            # Copy sheet mẫu và đổi tên
            target_sheet = workbook.copy_worksheet(source_sheet)
            target_sheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"
            worksheet = target_sheet
        else:
            # Sử dụng sheet đầu tiên
            worksheet = source_sheet
            worksheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"

        # Điền tên đơn vị
        worksheet['A1'] = f"{don_vi.ten_don_vi}"

        # Điền ngày tháng năm hiện tại
        worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"

        # Điền thông tin Mục lục hồ sơ
        worksheet['A3'] = f"MỤC LỤC HỒ SƠ ({muc_luc_ho_so.muc_luc_so})"

        # Điền dữ liệu Hồ sơ
        row_num = 6  # Bắt đầu từ dòng 6
        for hoso in ho_so_list.filter(muc_luc_ho_so=muc_luc_ho_so):
            worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
            worksheet['B{}'.format(row_num)] = hoso.ho_so_so
            worksheet['C{}'.format(row_num)] = hoso.tieu_de
            worksheet['D{}'.format(row_num)] = f"{hoso.thoi_gian_bat_dau.strftime('%d/%m/%Y')} - {hoso.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if hoso.thoi_gian_ket_thuc else ''}"
            worksheet['E{}'.format(row_num)] = f"{hoso.so_luong_to}"
            # Thêm cột Thời hạn bảo quản
            if hoso.thoi_han_bao_quan.ten == "Vĩnh viễn":
                worksheet['F{}'.format(row_num)] = hoso.ghi_chu
            else:
                worksheet['F{}'.format(row_num)] = hoso.thoi_han_bao_quan.ten 
                worksheet['G{}'.format(row_num)] = hoso.ghi_chu
            row_num += 1

    # Xóa sheet mẫu nếu nó vẫn còn sau khi duyệt
    if len(workbook.sheetnames) > len(muc_luc_ho_so_list):
        workbook.remove(source_sheet)

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="In_MLHS_{muc_luc_ho_so.muc_luc_so}.xlsx"'
    workbook.save(response)
    return response   

def in_muc_luc_van_ban_thuoc_ho_so(request):
    don_vi_id = request.GET.get("don_vi_id")
    ho_so_id = request.GET.get("ho_so_id")
    if not ho_so_id or not don_vi_id:
        return JsonResponse({"error": "Thiếu tham số ID hồ sơ"}, status=400)

    try:
        ho_so = HoSo.objects.get(pk=ho_so_id)
        don_vi = DonVi.objects.get(pk=don_vi_id)
        van_ban_list = VanBan.objects.filter(ho_so=ho_so)
    except HoSo.DoesNotExist:
        return JsonResponse({"error": "Hồ sơ không tồn tại"}, status=404)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)

    ho_so = get_object_or_404(HoSo, pk=ho_so_id)
    van_ban_list = VanBan.objects.filter(ho_so=ho_so)

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '5.MucLucVanBan-(6).xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year


    # Điền ngày tháng năm
    worksheet['E3'] = f"Đà nẵng, ngày {ngay} tháng {thang} năm {nam}"

    # Điền tên đơn vị
    worksheet['A1'] = f"{don_vi.ten_don_vi}"

    # Điền thông tin Hồ sơ
    worksheet['A4'] = f"MỤC LỤC VĂN BẢN (HỒ SƠ {ho_so.ho_so_so})"
    # Điền thông tin Hồ sơ
    worksheet['A5'] = f"Hồ sơ : {ho_so.tieu_de}"

    # Điền dữ liệu Văn bản
    row_num = 9  # Bắt đầu từ dòng 9
    for idx, van_ban in enumerate(van_ban_list):
        worksheet['A{}'.format(row_num)] = idx+1
        worksheet['B{}'.format(row_num)] = f"{van_ban.so_van_ban}/{van_ban.ky_hieu_cua_van_ban}"
        worksheet['C{}'.format(row_num)] = van_ban.ngay_ban_hanh.strftime('%d/%m/%Y')
        worksheet['D{}'.format(row_num)] = f"{van_ban.loai_van_ban} - {van_ban.trich_yeu_noi_dung}" if van_ban.loai_van_ban else van_ban.trich_yeu_noi_dung 
        worksheet['E{}'.format(row_num)] = van_ban.co_quan_to_chuc_ban_hanh
        worksheet['F{}'.format(row_num)] = f"{van_ban.tu_so}-{van_ban.den_so}" if van_ban.tu_so and van_ban.den_so else ''
        worksheet['G{}'.format(row_num)] = van_ban.ghi_chu
        row_num += 1

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="muc_luc_van_ban_{ho_so.ho_so_so}.xlsx"'
    workbook.save(response)
    return response

def in_bia_ho_so(request):
    don_vi_id = request.GET.get("don_vi_id")
    ho_so_id = request.GET.get("ho_so_id")
    phong_so = request.GET.get("phong_so")
    if not ho_so_id or not phong_so or not don_vi_id:
        return JsonResponse(
            {"error": "Thiếu tham số ID hồ sơ hoặc phông số"}, status=400
        )

    try:
        ho_so = HoSo.objects.get(pk=ho_so_id)
        phong_luu_tru = PhongLuuTru.objects.get(phong_so=phong_so)
        don_vi = DonVi.objects.get(pk=don_vi_id)
    except HoSo.DoesNotExist:
        return JsonResponse({"error": "Hồ sơ không tồn tại"}, status=404)
    except PhongLuuTru.DoesNotExist:
        return JsonResponse({"error": "Phông lưu trữ không tồn tại"}, status=404)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)

    ho_so = get_object_or_404(HoSo, pk=ho_so_id)

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '7.bia.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Điền dữ liệu
    worksheet['B3'] = f"{str(ho_so.phong_luu_tru.don_vi.ten_don_vi).upper()}"
    worksheet['B4'] = f"{ho_so.phong_luu_tru.name}"  # Giả sử tên trung tâm luôn cố định
    worksheet['C9'] = ho_so.ma_ho_so
    # worksheet['A14'] = "HỒ SƠ"
    worksheet['A15'] = ho_so.tieu_de
    worksheet['D18'] = f"Từ ngày: {ho_so.thoi_gian_bat_dau.strftime('%d/%m/%Y')}"
    worksheet['F18'] = f"Đến ngày: {ho_so.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if ho_so.thoi_gian_ket_thuc else ''}"
    worksheet['B19'] = f"Gồm: {ho_so.so_luong_to} tờ"
    # worksheet['C23'] = "Phông số:"
    worksheet['C22'] = ho_so.phong_luu_tru.phong_so
    # worksheet['C24'] = "Mục lục số:"
    worksheet['C23'] = ho_so.muc_luc_ho_so.muc_luc_so if ho_so.muc_luc_ho_so else ''
    # worksheet['C25'] = "Hồ sơ số:"
    worksheet['C24'] = ho_so.ho_so_so
    # worksheet['F23'] = "THỜI HẠN BẢO QUẢN"
    # worksheet['F23'] = f"{ho_so.thoi_han_bao_quan.so_nam} NĂM" if ho_so.thoi_han_bao_quan else ''
    worksheet['G23'] = ho_so.thoi_han_bao_quan.ten if ho_so.thoi_han_bao_quan else ''

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="In_bia_ho_so_{ho_so.ho_so_so}.xlsx"'
    workbook.save(response)
    return response

def in_phieu_muon(request):
    don_vi_id = request.GET.get("don_vi_id")
    so_phieu = request.GET.get("so_phieu")
    ho_ten = request.GET.get("ho_ten")
    so_cccd = request.GET.get("so_cccd")
    if not so_phieu or not ho_ten or not so_cccd or not don_vi_id:
        return JsonResponse(
            {"error": "Thiếu tham số số phiếu, họ tên hoặc số CCCD"}, status=400
        )

    try:
        don_vi = DonVi.objects.get(pk=don_vi_id)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '9.phieu-muon-tl.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Điền dữ liệu
    worksheet['B1'] = f"{don_vi.ten_don_vi}"
    worksheet['C3'] = f"Số: {so_phieu}"  # Giả sử tên trung tâm luôn cố định
    worksheet['C5'] = f"{ho_ten}"
    worksheet['C6'] = f"{so_cccd}"
    
    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="In_Phieu_muon.xlsx"'
    workbook.save(response)
    return response

def in_bao_cao_thong_ke(request):
    don_vi_id = request.GET.get("don_vi_id")
    ten_co_quan = request.GET.get("ten_co_quan")


    try:
        don_vi = DonVi.objects.get(pk=don_vi_id)
    except DonVi.DoesNotExist:
        return JsonResponse({"error": "Đơn vị không tồn tại"}, status=404)


    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'thongke', '8.bao cao thong ke.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    template_path = full_template_path
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Điền ngày tháng năm hiện tại
    worksheet['C22'] = f"Đà nẵng, ngày {ngay} tháng {thang} năm {nam}"

    # Thống kê dữ liệu
    tong_so_phong = PhongLuuTru.objects.filter(don_vi=don_vi).count()
    tong_so_mlhs = MucLucHoSo.objects.filter(phong_luu_tru__don_vi=don_vi).count()
    tong_so_ho_so = HoSo.objects.filter(phong_luu_tru__don_vi=don_vi).count()
    tong_so_van_ban = VanBan.objects.filter(ho_so__phong_luu_tru__don_vi=don_vi).count()
    tong_so_van_ban_da_so_hoa = VanBan.objects.filter(trang_thai='da_so_hoa', ho_so__phong_luu_tru__don_vi=don_vi).count()

    thoi_han_bao_quan_list = ThoiHanLuuTru.objects.all()
    thong_ke_ho_so_theo_thoi_han = {}
    for thoi_han in thoi_han_bao_quan_list:
        so_luong = HoSo.objects.filter(thoi_han_bao_quan=thoi_han, phong_luu_tru__don_vi=don_vi).count()
        thong_ke_ho_so_theo_thoi_han[thoi_han.ten] = so_luong

    # Điền tên đơn vị
    worksheet['A1'] = f"{don_vi.ten_don_vi}"

    # Điền dữ liệu thống kê vào Excel
    worksheet['C5'] = tong_so_phong
    worksheet['C6'] = tong_so_mlhs
    worksheet['C7'] = tong_so_ho_so
    worksheet['C8'] = tong_so_van_ban
    worksheet['C9'] = tong_so_van_ban_da_so_hoa

    # Thống kê hồ sơ theo thời hạn lưu trữ
    thoi_han_bao_quan_list = ThoiHanLuuTru.objects.all()
    row_num = 10
    for thoi_han in thoi_han_bao_quan_list:
        worksheet['B{}'.format(row_num)] = thoi_han.ten  # Điền tên thời hạn vào cột B
        so_luong = HoSo.objects.filter(
            thoi_han_bao_quan=thoi_han, 
            phong_luu_tru__don_vi=don_vi
        ).count()
        worksheet['C{}'.format(row_num)] = so_luong  # Điền số lượng vào cột C
        row_num += 1

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bao_cao_thong_ke_tai_lieu.xlsx"'
    workbook.save(response)
    return response

# View để lấy danh sách liên quan
def lay_danh_sach_lien_quan(request):
    don_vi_id = request.GET.get('don_vi_id')
    if don_vi_id:
        try:
            don_vi = DonVi.objects.get(pk=don_vi_id)

            # Lấy danh sách hồ sơ thuộc đơn vị
            ho_so_list = HoSo.objects.filter(phong_luu_tru__don_vi=don_vi)
            # Sắp xếp bằng Python, chuyển chuỗi thành số trước khi sắp xếp
            ho_so_list = sorted(ho_so_list, key=lambda obj: int(obj.ho_so_so))

            # Lấy danh sách phông lưu trữ thuộc đơn vị
            phong_luu_tru_list = PhongLuuTru.objects.filter(don_vi=don_vi)

            # Lấy danh sách thời hạn bảo quản thuộc đơn vị (distinct)
            thoi_han_bao_quan_list = ThoiHanLuuTru.objects.filter(
                hoso__phong_luu_tru__don_vi=don_vi
            ).distinct()

            # Lấy danh sách mục lục hồ sơ thuộc đơn vị
            muc_luc_ho_so_list = MucLucHoSo.objects.filter(
                phong_luu_tru__don_vi=don_vi
            )

            # Chuyển đổi dữ liệu sang JSON
            ho_so_data = [{'id': ho_so.id, 'ho_so_so': ho_so.ho_so_so} for ho_so in ho_so_list]
            phong_luu_tru_data = [{'id': phong.id, 'name': phong.name} for phong in phong_luu_tru_list]
            thoi_han_bao_quan_data = [{'id': thoi_han.id, 'ten': thoi_han.ten} for thoi_han in thoi_han_bao_quan_list]
            muc_luc_ho_so_data = [{'id': muc_luc.id, 'muc_luc_so': muc_luc.muc_luc_so} for muc_luc in muc_luc_ho_so_list]

            return JsonResponse({
                'ho_so': ho_so_data,
                'phong_luu_tru': phong_luu_tru_data,
                'thoi_han_bao_quan': thoi_han_bao_quan_data,
                'muc_luc_ho_so': muc_luc_ho_so_data,
            })

        except DonVi.DoesNotExist:
            return JsonResponse({'error': 'Đơn vị không tồn tại'}, status=404)
    else:
        return JsonResponse({'error': 'Thiếu tham số đơn vị'}, status=400)
    
@login_required
def export_mucluchoso_excel_vv(request, phong_luu_tru_id):
    phong_luu_tru = get_object_or_404(PhongLuuTru, pk=phong_luu_tru_id)
    don_vi = phong_luu_tru.don_vi

    # Lấy thời hạn bảo quản "Vĩnh viễn"
    thoi_han_vinh_vien = ThoiHanLuuTru.objects.get(ten="Vĩnh viễn")

    # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và có thời hạn bảo quản "Vĩnh viễn"
    ho_so_list = HoSo.objects.filter(
        phong_luu_tru=phong_luu_tru,
        phong_luu_tru__don_vi=don_vi,
        thoi_han_bao_quan=thoi_han_vinh_vien
    )

    # Lấy danh sách tất cả Mục lục hồ sơ có chứa Hồ sơ "Vĩnh viễn"
    muc_luc_ho_so_list = MucLucHoSo.objects.filter(
        id__in=ho_so_list.values_list('muc_luc_ho_so', flat=True).distinct()
    )

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs.xlsx')
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    workbook = openpyxl.load_workbook(full_template_path)

    # Lấy sheet đầu tiên làm sheet mẫu
    source_sheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Duyệt qua danh sách Mục lục hồ sơ
    for i, muc_luc_ho_so in enumerate(muc_luc_ho_so_list):
        if i > 0:
            # Copy sheet mẫu và đổi tên
            target_sheet = workbook.copy_worksheet(source_sheet)
            target_sheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"
            worksheet = target_sheet
        else:
            # Sử dụng sheet đầu tiên
            worksheet = source_sheet
            worksheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"

        # Điền tên đơn vị
        worksheet['A1'] = f"{don_vi.ten_don_vi}"

        # Điền ngày tháng năm hiện tại
        worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"

        # Điền thông tin Mục lục hồ sơ
        worksheet['A3'] = f"MỤC LỤC HỒ SƠ ({muc_luc_ho_so.muc_luc_so})"

        # Điền dữ liệu Hồ sơ
        row_num = 6  # Bắt đầu từ dòng 6
        for hoso in ho_so_list.filter(muc_luc_ho_so=muc_luc_ho_so):
            worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
            worksheet['B{}'.format(row_num)] = hoso.ho_so_so
            worksheet['C{}'.format(row_num)] = hoso.tieu_de
            worksheet['D{}'.format(row_num)] = f"{hoso.thoi_gian_bat_dau.strftime('%d/%m/%Y')} - {hoso.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if hoso.thoi_gian_ket_thuc else ''}"
            worksheet['E{}'.format(row_num)] = f"{hoso.so_luong_to}"
            worksheet['F{}'.format(row_num)] = hoso.ghi_chu
            row_num += 1

    # Xóa sheet mẫu nếu nó vẫn còn sau khi duyệt
    if len(workbook.sheetnames) > len(muc_luc_ho_so_list):
        workbook.remove(source_sheet)

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="danh_sach_ho_so_phong_{phong_luu_tru.phong_so}.xlsx"'
    workbook.save(response)
    return response

@login_required
def export_mucluchoso_excel_cth(request, phong_luu_tru_id):
    phong_luu_tru = get_object_or_404(PhongLuuTru, pk=phong_luu_tru_id)
    don_vi = phong_luu_tru.don_vi

    # Lấy thời hạn bảo quản "Vĩnh viễn"
    thoi_han_vinh_vien = ThoiHanLuuTru.objects.get(ten="Vĩnh viễn")

    # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và KHÔNG có thời hạn bảo quản "Vĩnh viễn"
    ho_so_list = HoSo.objects.filter(
        phong_luu_tru=phong_luu_tru,
        phong_luu_tru__don_vi=don_vi
    ).exclude(thoi_han_bao_quan=thoi_han_vinh_vien)

    # Lấy danh sách tất cả Mục lục hồ sơ có chứa Hồ sơ "Có thời hạn"
    muc_luc_ho_so_list = MucLucHoSo.objects.filter(
        id__in=ho_so_list.values_list('muc_luc_ho_so', flat=True).distinct()
    )

    # Load mẫu Excel
    template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs-cth.xlsx') # Thay đổi tên template
    full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
    workbook = openpyxl.load_workbook(full_template_path)

    # Lấy sheet đầu tiên làm sheet mẫu
    source_sheet = workbook.active

    # Lấy ngày tháng năm hiện tại
    now = datetime.now()
    ngay = now.day
    thang = now.month
    nam = now.year

    # Duyệt qua danh sách Mục lục hồ sơ
    for i, muc_luc_ho_so in enumerate(muc_luc_ho_so_list):
        if i > 0:
            # Copy sheet mẫu và đổi tên
            target_sheet = workbook.copy_worksheet(source_sheet)
            target_sheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"
            worksheet = target_sheet
        else:
            # Sử dụng sheet đầu tiên
            worksheet = source_sheet
            worksheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"

        # Điền tên đơn vị
        worksheet['A1'] = f"{don_vi.ten_don_vi}"

        # Điền ngày tháng năm hiện tại
        worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"

        # Điền thông tin Mục lục hồ sơ
        worksheet['A3'] = f"MỤC LỤC HỒ SƠ ({muc_luc_ho_so.muc_luc_so})"

        # Điền dữ liệu Hồ sơ
        row_num = 6  # Bắt đầu từ dòng 6
        for hoso in ho_so_list.filter(muc_luc_ho_so=muc_luc_ho_so):
            worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
            worksheet['B{}'.format(row_num)] = hoso.ho_so_so
            worksheet['C{}'.format(row_num)] = hoso.tieu_de
            worksheet['D{}'.format(row_num)] = f"{hoso.thoi_gian_bat_dau.strftime('%d/%m/%Y')} - {hoso.thoi_gian_ket_thuc.strftime('%d/%m/%Y') if hoso.thoi_gian_ket_thuc else ''}"
            worksheet['E{}'.format(row_num)] = f"{hoso.so_luong_to}"
            # Thêm cột Thời hạn bảo quản
            worksheet['F{}'.format(row_num)] = hoso.thoi_han_bao_quan.ten 
            worksheet['G{}'.format(row_num)] = hoso.ghi_chu
            row_num += 1

    # Xóa sheet mẫu nếu nó vẫn còn sau khi duyệt
    if len(workbook.sheetnames) > len(muc_luc_ho_so_list):
        workbook.remove(source_sheet)

    # Tạo response HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="danh_sach_ho_so_phong_{phong_luu_tru.phong_so}.xlsx"'
    workbook.save(response)
    return response

@login_required
def export_mucluchoso_excel_loaibo(request, phong_luu_tru_id):
    from datetime import timedelta
    phong_luu_tru = get_object_or_404(PhongLuuTru, pk=phong_luu_tru_id)
    don_vi = phong_luu_tru.don_vi
    try:
        # Lấy danh sách tất cả Hồ sơ thuộc PhongLuuTru và có thời hạn bảo quản
        ho_so_co_thoi_han = HoSo.objects.filter(
            phong_luu_tru=phong_luu_tru,
            phong_luu_tru__don_vi=don_vi
        ).exclude(thoi_han_bao_quan__ten="Vĩnh viễn")

        # Lấy ngày hiện tại
        now = datetime.now().date()

        # Lọc danh sách hồ sơ hết hạn
        ho_so_het_han = []
        for ho_so in ho_so_co_thoi_han:
            thoi_gian_het_han = ho_so.thoi_gian_bat_dau + timedelta(days=ho_so.thoi_han_bao_quan.so_nam * 365.25)
            if thoi_gian_het_han < now:
                ho_so_het_han.append(ho_so)

        # Lấy danh sách tất cả Mục lục hồ sơ có chứa Hồ sơ hết hạn
        muc_luc_ho_so_list = MucLucHoSo.objects.filter(
            id__in=[ho_so.muc_luc_ho_so_id for ho_so in ho_so_het_han]
        ).distinct()

        # Load mẫu Excel
        template_path = os.path.join('admin', 'home', 'mucluchoso', '4.mau-mlhs-loaibo.xlsx') # Thay đổi tên template
        full_template_path = os.path.join(settings.BASE_DIR, 'home','templates', template_path)
        workbook = openpyxl.load_workbook(full_template_path)

        # Lấy sheet đầu tiên làm sheet mẫu
        source_sheet = workbook.active

        # Lấy ngày tháng năm hiện tại
        ngay = now.day
        thang = now.month
        nam = now.year

        # Duyệt qua danh sách Mục lục hồ sơ
        for i, muc_luc_ho_so in enumerate(muc_luc_ho_so_list):
            if i > 0:
                # Copy sheet mẫu và đổi tên
                target_sheet = workbook.copy_worksheet(source_sheet)
                target_sheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"
                worksheet = target_sheet
            else:
                # Sử dụng sheet đầu tiên
                worksheet = source_sheet
                worksheet.title = f"Mục lục {muc_luc_ho_so.muc_luc_so}"

            # Điền tên đơn vị
            worksheet['A1'] = f"{don_vi.ten_don_vi}"

            # Điền ngày tháng năm hiện tại
            worksheet['D1'] = f"Ngày {ngay} tháng {thang} năm {nam}"

            # Điền thông tin Mục lục hồ sơ
            worksheet['A3'] = f"DANH MỤC TÀI LIỆU HẾT GIÁ TRỊ ({muc_luc_ho_so.muc_luc_so})"

            # Điền dữ liệu Hồ sơ
            row_num = 6  # Bắt đầu từ dòng 6
            for hoso in ho_so_het_han:
                if hoso.muc_luc_ho_so == muc_luc_ho_so:
                    worksheet['A{}'.format(row_num)] = hoso.hop_luu_tru.hop_so if hoso.hop_luu_tru else ''
                    worksheet['B{}'.format(row_num)] = hoso.ho_so_so
                    worksheet['C{}'.format(row_num)] = hoso.tieu_de
                    worksheet['D{}'.format(row_num)] = f"{hoso.thoi_han_bao_quan.ten}\n Hết giá trị"
                    worksheet['E{}'.format(row_num)] = hoso.ghi_chu
                    # Thêm cột Lý do loại

                    row_num += 1

        # Xóa sheet mẫu nếu nó vẫn còn sau khi duyệt
        if len(workbook.sheetnames) > len(muc_luc_ho_so_list):
            workbook.remove(source_sheet)

        # Tạo response HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="danh_sach_ho_so_phong_{phong_luu_tru.phong_so}_loaibos.xlsx"'
        workbook.save(response)
        return response
    except:
        return JsonResponse({'Info': 'Không có tài liệu hết hạn'}, status=200)